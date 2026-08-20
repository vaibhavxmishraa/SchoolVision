import io
from datetime import date
import pandas as pd
from sqlalchemy.orm import Session
from .models import Student, Attendance

       
def build_report(db: Session, class_section: str = None) -> io.BytesIO:
    q = db.query(Student)
    if class_section:
        q = q.filter(Student.class_section == class_section)
    students = q.all()

    rows = []
    for s in students:
        records = db.query(Attendance).filter(Attendance.student_id == s.id).all()

        # Per-day aggregation
        days = {}
        for r in records:
            d = r.date
            days.setdefault(d, {"in": "-", "out": "-"})
            if r.session_type == "Check-In":
                days[d]["in"] = r.status
            else:
                days[d]["out"] = r.status

        total_days = len(days) if days else 1
        present_marks = sum(
            1 for v in days.values()
            for x in (v["in"], v["out"]) if x in ("Present", "Manual Override")
        )
        pct = round((present_marks / (total_days * 2)) * 100, 1) if days else 0.0

        if days:
            for d, v in sorted(days.items()):
                rows.append({
                    "Date": d.strftime("%d-%m-%Y"),
                    "Roll Number": s.roll_number,
                    "Student Name": s.name,
                    "Class": s.class_section,
                    "Morning Check-In": v["in"],
                    "Evening Check-Out": v["out"],
                    "Total Attendance %": f"{pct}%",
                })
        else:
            rows.append({
                "Date": date.today().strftime("%d-%m-%Y"),
                "Roll Number": s.roll_number, "Student Name": s.name,
                "Class": s.class_section, "Morning Check-In": "Absent",
                "Evening Check-Out": "Absent", "Total Attendance %": "0%",
            })

    df = pd.DataFrame(rows) if rows else pd.DataFrame(columns=[
        "Date", "Roll Number", "Student Name", "Class",
        "Morning Check-In", "Evening Check-Out", "Total Attendance %"])

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Attendance")
        ws = writer.sheets["Attendance"]
        for col in ws.columns:
            width = max((len(str(c.value)) for c in col if c.value), default=10) + 4
            ws.column_dimensions[col[0].column_letter].width = width
        # Header styling
        from openpyxl.styles import Font, PatternFill, Alignment
        fill = PatternFill("solid", fgColor="1E3A8A")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")
    buffer.seek(0)
    return buffer